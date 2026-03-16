from __future__ import annotations

import csv
import io
from datetime import datetime, timezone, date
from typing import Optional, Dict, Any, List, Tuple

from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert

from celery_app import celery_app
from app.database import sync_session_maker
from app.models.import_job import ImportJob
from app.models.transaction import Transaction
from app.models.notification import Notification
from app.services.storage_s3 import get_bytes, S3ObjectRef, sha256_bytes


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _parse_dt(value: str) -> Optional[datetime]:
    v = (value or "").strip()
    if not v:
        return None
    if v.endswith("Z"):
        v = v[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(v)
    except Exception:
        return None


def _norm_key(k: Optional[str]) -> Optional[str]:
    if k is None:
        return None
    return k.replace("\ufeff", "").strip()


def _fix_mojibake(s: str) -> str:
    if not s:
        return s
    if ("Ð" not in s) and ("Ñ" not in s):
        return s
    try:
        return s.encode("latin-1").decode("utf-8")
    except Exception:
        return s


def _norm_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return _fix_mojibake(s)


def _norm_row(row: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in row.items():
        nk = _norm_key(k)
        if nk:
            out[nk] = v
    return out


def _decode_csv_bytes(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


def _compute_fingerprint(
    tenant_id: str,
    bank_account_id: str,
    occurred_at_iso: str,
    amount: str,
    ttype: str,
    inn: str,
    desc: str,
    doc: str,
) -> str:
    raw = "|".join(
        [
            tenant_id or "",
            bank_account_id or "",
            occurred_at_iso or "",
            amount or "",
            ttype or "",
            inn or "",
            desc or "",
            doc or "",
        ]
    )
    return sha256_bytes(raw.encode("utf-8"))


def _notify(db, job: ImportJob, title: str, message: str, ntype: str, payload: Optional[dict] = None):
    if not job.user_id:
        return
    db.add(
        Notification(
            tenant_id=job.tenant_id,
            user_id=job.user_id,
            title=title,
            message=message,
            notification_type=ntype,
            payload=payload or {},
            related_type="import_job",
            related_id=str(job.id),
        )
    )


def _is_xlsx(job: ImportJob, data: bytes) -> bool:
    fn = (job.file_name or "").lower()
    mime = (job.file_mime or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return True
    if "spreadsheetml" in mime or "excel" in mime:
        return True
    if len(data) >= 4 and data[0:4] == b"PK\x03\x04":
        return True
    return False


# ---------------- XLSX helpers ----------------

def _xlsx_find_header(ws) -> Tuple[int, List[str]]:
    """
    Header row detection:
      must contain 'дата' and one of:
        - 'дебет'+'кредит' (PSB)
        - 'сумма'
        - 'поступление'+'списание' (Sber)
    """
    def cell_str(x) -> str:
        return (str(x).strip() if x is not None else "")

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
        vals = [cell_str(v) for v in row]
        joined = " ".join(vals).lower()

        has_date = "дата" in joined
        has_debit_credit = ("дебет" in joined and "кредит" in joined)
        has_amount = "сумма" in joined
        has_in_out = ("поступление" in joined and "списание" in joined)

        if has_date and (has_debit_credit or has_amount or has_in_out):
            return i, vals

    raise ValueError("XLSX: header row not found (expected 'Дата' + ('Дебет/Кредит' or 'Сумма' or 'Поступление/Списание'))")


def _xlsx_guess_col(headers: List[str], contains_any: List[str]) -> Optional[int]:
    hs = [h.strip().lower() for h in headers]
    keys = [k.lower() for k in contains_any]
    for i, h in enumerate(hs):
        if not h:
            continue
        for k in keys:
            if k in h:
                return i
    return None


def _xlsx_skip_row(date_cell: Any) -> bool:
    if date_cell is None:
        return True
    if isinstance(date_cell, str):
        low = date_cell.lower()
        if "сальдо" in low or "итог" in low:
            return True
    return False


def _to_dt_utc(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=timezone.utc)
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day, tzinfo=timezone.utc)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            dt = datetime.strptime(s.split()[0], "%d.%m.%Y")
            return dt.replace(tzinfo=timezone.utc)
        except Exception:
            return None
    return None


def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _norm_text(v)
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _xlsx_to_rows_generic(data: bytes) -> List[Dict[str, Any]]:
    """
    Universal XLSX parser (PSB + Sber):
    Outputs dicts compatible with CSV pipeline keys:
      occurred_at, transaction_type, amount, currency, description, counterparty_name, counterparty_inn, document_number, external_id
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx, headers = _xlsx_find_header(ws)

    i_date = _xlsx_guess_col(headers, ["дата"])

    i_debit = _xlsx_guess_col(headers, ["дебет"])
    i_credit = _xlsx_guess_col(headers, ["кредит"])

    i_incoming = _xlsx_guess_col(headers, ["поступление", "зачисление", "приход"])
    i_outgoing = _xlsx_guess_col(headers, ["списание", "расход"])

    i_amount = _xlsx_guess_col(headers, ["сумма"])

    i_currency_1 = _xlsx_guess_col(headers, ["валюта"])
    i_currency_2 = None
    if i_currency_1 is not None:
        hs = [h.strip().lower() for h in headers]
        for j in range(i_currency_1 + 1, len(hs)):
            if hs[j] == hs[i_currency_1] and hs[j] in ("валюта", "currency"):
                i_currency_2 = j
                break

    i_desc = _xlsx_guess_col(headers, ["назначение", "назначение платежа", "описание", "детали"])
    i_counterparty = _xlsx_guess_col(headers, ["контрагент"])
    i_inn = _xlsx_guess_col(headers, ["инн"])
    i_doc = _xlsx_guess_col(headers, ["док", "номер документа", "номер"])

    out: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        v_date = row[i_date] if i_date is not None and i_date < len(row) else None
        if _xlsx_skip_row(v_date):
            continue

        occurred_at_dt = _to_dt_utc(v_date)
        if occurred_at_dt is None:
            continue

        ttype = None
        amount_val = None

        if i_incoming is not None and i_outgoing is not None:
            inc = _num(row[i_incoming] if i_incoming < len(row) else None)
            outv = _num(row[i_outgoing] if i_outgoing < len(row) else None)
            if inc and (not outv or outv == 0):
                ttype = "incoming"
                amount_val = abs(inc)
            elif outv and (not inc or inc == 0):
                ttype = "outgoing"
                amount_val = abs(outv)
            else:
                continue

        elif i_debit is not None and i_credit is not None:
            d = _num(row[i_debit] if i_debit < len(row) else None)
            c = _num(row[i_credit] if i_credit < len(row) else None)
            if d and (not c or c == 0):
                ttype = "outgoing"
                amount_val = abs(d)
            elif c and (not d or d == 0):
                ttype = "incoming"
                amount_val = abs(c)
            else:
                continue

        elif i_amount is not None:
            a = _num(row[i_amount] if i_amount < len(row) else None)
            if a is None:
                continue
            ttype = "incoming" if a >= 0 else "outgoing"
            amount_val = abs(a)

        else:
            raise ValueError("XLSX: cannot determine amount columns")

        currency = "RUB"
        if ttype == "incoming":
            if i_currency_1 is not None and i_currency_1 < len(row):
                currency = _norm_text(row[i_currency_1]) or currency
        else:
            if i_currency_2 is not None and i_currency_2 < len(row):
                currency = _norm_text(row[i_currency_2]) or currency
            elif i_currency_1 is not None and i_currency_1 < len(row):
                currency = _norm_text(row[i_currency_1]) or currency

        desc = ""
        if i_desc is not None and i_desc < len(row):
            desc = _norm_text(row[i_desc]) or ""

        counterparty = None
        if i_counterparty is not None and i_counterparty < len(row):
            counterparty = _norm_text(row[i_counterparty])

        inn = ""
        if i_inn is not None and i_inn < len(row):
            inn = _norm_text(row[i_inn]) or ""

        doc = None
        if i_doc is not None and i_doc < len(row):
            doc = _norm_text(row[i_doc])

        out.append(
            {
                "occurred_at": occurred_at_dt.isoformat(),
                "transaction_type": ttype,
                "amount": str(amount_val).replace(",", "."),
                "currency": currency,
                "description": desc,
                "counterparty_name": counterparty,
                "counterparty_inn": inn,
                "document_number": doc,
                "external_id": doc,
            }
        )

    if not out:
        raise ValueError("XLSX: parsed 0 operations (check headers/format)")
    return out


# ---------------- Celery task ----------------

@celery_app.task(bind=True, name="app.tasks.bank_import.import_bank_statement_job", max_retries=3)
def import_bank_statement_job(self, job_id: str):
    with sync_session_maker() as db:
        job = db.execute(select(ImportJob).where(ImportJob.id == job_id)).scalar_one_or_none()
        if not job:
            return {"status": "error", "message": "job not found"}

        if job.status == "SUCCESS":
            return {"status": "skipped", "message": "job already SUCCESS"}

        bank_account_id = (job.meta or {}).get("bank_account_id")
        if not bank_account_id:
            job.status = "FAILED"
            job.error = "bank_account_id is required in job.meta"
            job.finished_at = _utcnow()
            _notify(db, job, "Импорт выписки: ошибка", job.error, "error", payload={"job_id": job_id})
            db.commit()
            return {"status": "failed", "message": job.error}

        job.status = "PROCESSING"
        job.started_at = _utcnow()
        job.error = None
        db.commit()

        try:
            data = get_bytes(S3ObjectRef(bucket=job.s3_bucket, key=job.s3_key))

            if _is_xlsx(job, data):
                parsed_rows = _xlsx_to_rows_generic(data)
                reader_rows = [_norm_row({k: v for k, v in r.items()}) for r in parsed_rows]
            else:
                text = _decode_csv_bytes(data)
                if text.startswith("\ufeff"):
                    text = text[1:]
                reader = csv.DictReader(io.StringIO(text))
                reader_rows = []
                for raw in reader:
                    reader_rows.append(_norm_row(raw))

            rows = []
            total = 0

            for r in reader_rows:
                total += 1

                occurred_at_raw = _norm_text(r.get("occurred_at") or r.get("date") or "") or ""
                occurred_at_dt = _parse_dt(occurred_at_raw)
                if occurred_at_dt is None:
                    raise ValueError(f"Invalid occurred_at at row {total}: '{occurred_at_raw}' (keys={list(r.keys())})")

                ttype = (_norm_text(r.get("transaction_type") or r.get("type")) or "").lower()
                amount_raw = (_norm_text(r.get("amount")) or "0").replace(",", ".")
                currency = _norm_text(r.get("currency")) or "RUB"

                inn = _norm_text(r.get("counterparty_inn")) or ""
                name = _norm_text(r.get("counterparty_name"))
                acc = _norm_text(r.get("counterparty_account"))
                desc = _norm_text(r.get("description")) or ""
                doc = _norm_text(r.get("document_number"))
                external_id = _norm_text(r.get("external_id"))

                if ttype not in ("incoming", "outgoing"):
                    try:
                        v = float(amount_raw)
                        ttype = "incoming" if v >= 0 else "outgoing"
                        amount_raw = str(abs(v))
                    except Exception:
                        ttype = "outgoing"

                fp = _compute_fingerprint(
                    job.tenant_id,
                    bank_account_id,
                    occurred_at_dt.isoformat(),
                    str(amount_raw),
                    ttype,
                    inn,
                    desc,
                    doc or "",
                )

                rows.append(
                    dict(
                        tenant_id=job.tenant_id,
                        bank_account_id=bank_account_id,
                        import_job_id=str(job.id),
                        transaction_type=ttype,
                        amount=amount_raw,
                        currency=currency,
                        occurred_at=occurred_at_dt,
                        counterparty_name=name,
                        counterparty_inn=inn or None,
                        counterparty_account=acc,
                        description=desc,
                        document_number=doc,
                        fingerprint=fp,
                        external_id=external_id,
                        is_processed=False,
                    )
                )

            job.total_rows = total

            inserted = 0
            with_external = [x for x in rows if x.get("external_id")]
            without_external = [x for x in rows if not x.get("external_id")]

            if with_external:
                stmt = insert(Transaction).values(with_external)
                stmt = stmt.on_conflict_do_nothing(
                    index_elements=["tenant_id", "bank_account_id", "external_id"],
                    index_where=Transaction.external_id.isnot(None),
                )
                res = db.execute(stmt)
                inserted += res.rowcount or 0

            if without_external:
                stmt = insert(Transaction).values(without_external)
                stmt = stmt.on_conflict_do_nothing(
                    index_elements=["tenant_id", "fingerprint"],
                    index_where=Transaction.fingerprint.isnot(None),
                )
                res = db.execute(stmt)
                inserted += res.rowcount or 0

            job.inserted_rows = inserted
            job.skipped_rows = total - inserted
            job.status = "SUCCESS"
            job.finished_at = _utcnow()

            _notify(
                db,
                job,
                "Импорт выписки: успешно",
                f"Файл '{job.file_name}' обработан. Добавлено: {inserted}, пропущено: {job.skipped_rows}.",
                "success",
                payload={"job_id": job_id, "inserted": inserted, "skipped": job.skipped_rows},
            )

            db.commit()
            return {"status": "success", "job_id": job_id, "inserted": inserted, "skipped": job.skipped_rows}

        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass

            job = db.execute(select(ImportJob).where(ImportJob.id == job_id)).scalar_one_or_none()
            if job:
                job.status = "FAILED"
                job.error = str(exc)
                job.finished_at = _utcnow()
                _notify(
                    db,
                    job,
                    "Импорт выписки: ошибка",
                    f"Файл '{job.file_name}' не обработан: {job.error}",
                    "error",
                    payload={"job_id": job_id, "error": job.error},
                )
                db.commit()

            raise self.retry(exc=exc, countdown=30)